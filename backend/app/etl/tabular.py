from collections.abc import Iterator
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from app.territorial import normalize_territory_name
from openpyxl import load_workbook

REQUIRED_COLUMNS = [
    "fecha",
    "hora",
    "id_delito",
    "distrito",
    "direccion",
    "latitud",
    "longitud",
    "fuente_registro",
    "descripcion",
]
OPTIONAL_COLUMNS = [
    "id_comisaria",
]

DATE_FORMATS = ["%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"]
TIME_FORMATS = ["%H:%M", "%H:%M:%S"]
DATE_FIELDS = {"fecha"}
TIME_FIELDS = {"hora"}


class TabularValidationError(ValueError):
    def __init__(
        self,
        message: str,
        *,
        code: str = "ERROR_VALIDACION",
        staging_payload: dict[str, Any] | None = None,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.staging_payload = staging_payload or {}


def normalize_header(value: object) -> str:
    return str(value).strip() if value is not None else ""


def normalize_cell(value: object, field_name: str = "") -> str:
    normalized_name = field_name.strip().lower()

    if value is None:
        return ""
    if isinstance(value, datetime):
        if normalized_name in DATE_FIELDS:
            return value.strftime("%Y-%m-%d")
        if normalized_name in TIME_FIELDS:
            return value.strftime("%H:%M:%S")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return str(value).strip()


def read_tabular_source(
    file_path: Path,
    sheet_name: str | None = None,
) -> tuple[list[str], Iterator[tuple[int, dict[str, str]]]]:
    suffix = file_path.suffix.lower()

    if suffix == ".csv":
        return _iter_csv_rows(file_path)

    if suffix == ".xlsx":
        return _iter_excel_rows(file_path, sheet_name=sheet_name)

    raise TabularValidationError(
        f"Formato no soportado: {file_path.suffix}. Usa un archivo .csv o .xlsx."
    )


def validate_required_columns(fieldnames: list[str] | None) -> None:
    if fieldnames is None:
        raise TabularValidationError("El archivo no contiene cabecera.")

    missing_columns = [column for column in REQUIRED_COLUMNS if column not in fieldnames]
    if missing_columns:
        raise TabularValidationError(
            f"Faltan columnas requeridas: {', '.join(missing_columns)}."
        )


def parse_date(value: str) -> date:
    raw = (value or "").strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise TabularValidationError(f"Fecha invalida: {value}")


def parse_time(value: str) -> time:
    raw = (value or "").strip()
    for fmt in TIME_FORMATS:
        try:
            return datetime.strptime(raw, fmt).time()
        except ValueError:
            continue
    raise TabularValidationError(f"Hora invalida: {value}")


def parse_float(value: str, field_name: str) -> float:
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise TabularValidationError(f"{field_name} invalido: {value}") from exc


def parse_int(value: str, field_name: str) -> int:
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise TabularValidationError(f"{field_name} invalido: {value}") from exc


def parse_optional_int(value: str, field_name: str) -> int | None:
    raw = (value or "").strip()
    if raw == "":
        return None
    return parse_int(raw, field_name)


def parse_optional_positive_int_relaxed(value: str) -> tuple[int | None, bool]:
    raw = (value or "").strip()
    if raw == "":
        return None, False

    try:
        parsed = int(raw)
    except (TypeError, ValueError):
        return None, True

    if parsed <= 0:
        return None, True

    return parsed, False


def clean_text(value: str, field_name: str, upper: bool = False) -> str:
    text = (value or "").strip()
    if not text:
        raise TabularValidationError(f"{field_name} vacio")
    return text.upper() if upper else text


def build_territorial_staging_payload(
    *,
    estado_territorial: str,
    motivo_territorial: str,
    id_comisaria_original: int | None = None,
) -> dict[str, Any]:
    return {
        "id_comisaria_original": id_comisaria_original,
        "id_comisaria_resuelta": None,
        "nombre_comisaria_resuelta": None,
        "estado_territorial": estado_territorial,
        "regla_territorial": "REVISION_MANUAL",
        "motivo_territorial": motivo_territorial,
        "conflicto_territorial": False,
    }


def parse_coordinates(
    latitud_value: str,
    longitud_value: str,
    *,
    id_comisaria_original: int | None,
) -> tuple[float, float]:
    latitud_raw = (latitud_value or "").strip()
    longitud_raw = (longitud_value or "").strip()
    missing_fields = [
        field_name
        for field_name, raw_value in (("latitud", latitud_raw), ("longitud", longitud_raw))
        if raw_value == ""
    ]

    if missing_fields:
        joined_fields = " y ".join(missing_fields)
        message = f"Coordenadas incompletas: falta {joined_fields}."
        raise TabularValidationError(
            message,
            staging_payload=build_territorial_staging_payload(
                estado_territorial="COORDENADAS_INCOMPLETAS",
                motivo_territorial=message,
                id_comisaria_original=id_comisaria_original,
            ),
        )

    try:
        latitud = float(latitud_raw)
    except (TypeError, ValueError) as exc:
        message = f"latitud invalida: {latitud_value}"
        raise TabularValidationError(
            message,
            staging_payload=build_territorial_staging_payload(
                estado_territorial="COORDENADAS_INVALIDAS",
                motivo_territorial=message,
                id_comisaria_original=id_comisaria_original,
            ),
        ) from exc

    try:
        longitud = float(longitud_raw)
    except (TypeError, ValueError) as exc:
        message = f"longitud invalida: {longitud_value}"
        raise TabularValidationError(
            message,
            staging_payload=build_territorial_staging_payload(
                estado_territorial="COORDENADAS_INVALIDAS",
                motivo_territorial=message,
                id_comisaria_original=id_comisaria_original,
            ),
        ) from exc

    if not (-90 <= latitud <= 90):
        message = f"latitud fuera de rango: {latitud}"
        raise TabularValidationError(
            message,
            staging_payload=build_territorial_staging_payload(
                estado_territorial="COORDENADAS_INVALIDAS",
                motivo_territorial=message,
                id_comisaria_original=id_comisaria_original,
            ),
        )

    if not (-180 <= longitud <= 180):
        message = f"longitud fuera de rango: {longitud}"
        raise TabularValidationError(
            message,
            staging_payload=build_territorial_staging_payload(
                estado_territorial="COORDENADAS_INVALIDAS",
                motivo_territorial=message,
                id_comisaria_original=id_comisaria_original,
            ),
        )

    return latitud, longitud


def validate_row(raw_row: dict[str, str]) -> dict[str, Any]:
    for column in REQUIRED_COLUMNS:
        if column not in raw_row:
            raise TabularValidationError(f"Falta columna requerida: {column}")

    id_comisaria_original, id_comisaria_original_invalido = parse_optional_positive_int_relaxed(
        raw_row.get("id_comisaria", "")
    )
    latitud, longitud = parse_coordinates(
        raw_row.get("latitud", ""),
        raw_row.get("longitud", ""),
        id_comisaria_original=id_comisaria_original,
    )
    fecha = parse_date(raw_row["fecha"])
    hora = parse_time(raw_row["hora"])
    id_delito = parse_int(raw_row["id_delito"], "id_delito")

    if id_delito <= 0:
        raise TabularValidationError("id_delito debe ser mayor que 0")

    return {
        "fecha": fecha,
        "hora": hora,
        "id_delito": id_delito,
        "distrito": normalize_territory_name(clean_text(raw_row["distrito"], "distrito")),
        "direccion": clean_text(raw_row["direccion"], "direccion"),
        "latitud": latitud,
        "longitud": longitud,
        "id_comisaria": None,
        "id_comisaria_original": id_comisaria_original,
        "id_comisaria_original_invalido": id_comisaria_original_invalido,
        "id_comisaria_resuelta": None,
        "nombre_comisaria_resuelta": None,
        "estado_territorial": "SIN_EVALUAR",
        "regla_territorial": None,
        "motivo_territorial": None,
        "conflicto_territorial": False,
        "fuente_registro": clean_text(raw_row["fuente_registro"], "fuente_registro"),
        "descripcion": clean_text(raw_row["descripcion"], "descripcion"),
    }


def build_row_fingerprint(clean_row: dict[str, Any]) -> tuple[object, ...]:
    return (
        clean_row["fecha"],
        clean_row["hora"],
        clean_row["id_delito"],
        clean_row["distrito"],
        clean_row["direccion"],
        round(float(clean_row["latitud"]), 6),
        round(float(clean_row["longitud"]), 6),
        clean_row["id_comisaria"],
        clean_row["fuente_registro"],
        clean_row["descripcion"],
    )


def register_row_fingerprint(
    seen_fingerprints: dict[tuple[object, ...], int],
    clean_row: dict[str, Any],
    line_number: int,
) -> int | None:
    fingerprint = build_row_fingerprint(clean_row)
    previous_line = seen_fingerprints.get(fingerprint)

    if previous_line is None:
        seen_fingerprints[fingerprint] = line_number

    return previous_line


def build_error_entry(
    line_number: int,
    code: str,
    message: str,
    raw_row: dict[str, str] | None = None,
    *,
    context: str | None = None,
) -> dict[str, Any]:
    return {
        "linea": line_number,
        "codigo": code,
        "error": message,
        "data": context or str(raw_row or {}),
    }


def _iter_csv_rows(file_path: Path) -> tuple[list[str], Iterator[tuple[int, dict[str, str]]]]:
    import csv

    handle = file_path.open("r", encoding="utf-8-sig", newline="")
    reader = csv.DictReader(handle)
    fieldnames = reader.fieldnames or []

    def generator() -> Iterator[tuple[int, dict[str, str]]]:
        try:
            for line_number, raw_row in enumerate(reader, start=2):
                normalized_row = {
                    key: normalize_cell(value, key)
                    for key, value in (raw_row or {}).items()
                    if key is not None
                }
                yield line_number, normalized_row
        finally:
            handle.close()

    return fieldnames, generator()


def _iter_excel_rows(
    file_path: Path,
    sheet_name: str | None = None,
) -> tuple[list[str], Iterator[tuple[int, dict[str, str]]]]:
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise TabularValidationError(f"La hoja '{sheet_name}' no existe en el archivo Excel.")
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook[workbook.sheetnames[0]]

    row_iterator = worksheet.iter_rows(values_only=True)
    header_row = next(row_iterator, None)
    fieldnames = [normalize_header(value) for value in header_row] if header_row else []

    def generator() -> Iterator[tuple[int, dict[str, str]]]:
        try:
            for line_number, row_values in enumerate(row_iterator, start=2):
                padded_values = list(row_values or ())
                if len(padded_values) < len(fieldnames):
                    padded_values.extend([None] * (len(fieldnames) - len(padded_values)))

                normalized_row = {
                    fieldnames[index]: normalize_cell(padded_values[index], fieldnames[index])
                    if index < len(padded_values)
                    else ""
                    for index in range(len(fieldnames))
                    if fieldnames[index]
                }
                yield line_number, normalized_row
        finally:
            workbook.close()

    return fieldnames, generator()
